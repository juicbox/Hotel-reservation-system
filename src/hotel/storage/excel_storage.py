from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from hotel.enums import (
    BookingStatus,
    RoomStatus,
    RoomType,
    ServiceCategory,
    ShiftType,
    UserRole,
)
from hotel.hotel_system import HotelSystem, build_user_from_role
from hotel.models import Booking, Guest, Room, Service, Staff


class ExcelStorage:
    """Persistence helper for Rooms, Users, Bookings, and Services."""

    def __init__(self, file_path: str | Path) -> None:
        self.file_path = Path(file_path)

    def save(self, system: HotelSystem) -> None:
        """Write the complete in-memory hotel system to an Excel workbook."""
        workbook = Workbook()
        workbook.remove(workbook.active)

        self._write_rooms(workbook, system)
        self._write_users(workbook, system)
        self._write_bookings(workbook, system)
        self._write_services_catalog(workbook, system)
        workbook.save(self.file_path)

    def load(self, system: HotelSystem) -> bool:
        """Load data into the system if the Excel file already exists."""
        if not self.file_path.exists():
            return False

        workbook = load_workbook(self.file_path)
        # Clear preload/session data before replacing it with persisted records.
        system.rooms.clear()
        system.users.clear()
        system.bookings.clear()
        system.services.clear()

        self._read_rooms(workbook, system)
        self._read_users(workbook, system)
        self._read_bookings(workbook, system)
        self._read_services_catalog(workbook, system)
        return True

    def _write_rooms(self, workbook: Workbook, system: HotelSystem) -> None:
        """Persist room inventory and current room status."""
        sheet = workbook.create_sheet("rooms")
        sheet.append(
            [
                "room_id",
                "room_number",
                "room_type",
                "floor",
                "capacity",
                "base_nightly_rate",
                "current_status",
                "seasonal_multiplier",
                "amenities",
            ]
        )
        for room in system.rooms.values():
            sheet.append(
                [
                    room.room_id,
                    room.room_number,
                    room.room_type.value,
                    room.floor,
                    room.capacity,
                    room.base_nightly_rate,
                    room.current_status.value,
                    room.seasonal_multiplier,
                    ",".join(room.amenities),
                ]
            )

    def _write_users(self, workbook: Workbook, system: HotelSystem) -> None:
        """Persist users, including guest loyalty points and staff access level."""
        sheet = workbook.create_sheet("users")
        sheet.append(
            [
                "user_id",
                "full_name",
                "email",
                "password_hash",
                "phone",
                "role",
                "created_at",
                "is_active",
                "access_level",
                "loyalty_points",
            ]
        )
        for user in system.users.values():
            access_level = user.access_level if isinstance(user, Staff) else ""
            loyalty_points = user.loyalty_points if isinstance(user, Guest) else ""
            sheet.append(
                [
                    user.user_id,
                    user.full_name,
                    user.email,
                    user.password_hash,
                    user.phone,
                    user.role.value,
                    user.created_at.isoformat(),
                    int(user.is_active),
                    access_level,
                    loyalty_points,
                ]
            )

    def _write_bookings(self, workbook: Workbook, system: HotelSystem) -> None:
        """Persist bookings and serialize attached services into one cell."""
        sheet = workbook.create_sheet("bookings")
        sheet.append(
            [
                "booking_id",
                "guest_id",
                "room_id",
                "check_in_date",
                "check_out_date",
                "status",
                "total_charges",
                "deposit_paid",
                "notes",
                "created_by",
                "modified_at",
                "services",
            ]
        )
        for booking in system.bookings.values():
            sheet.append(
                [
                    booking.booking_id,
                    booking.guest_id,
                    booking.room_id,
                    booking.check_in_date.isoformat(),
                    booking.check_out_date.isoformat(),
                    booking.status.value,
                    booking.total_charges,
                    booking.deposit_paid,
                    booking.notes,
                    booking.created_by,
                    booking.modified_at.isoformat(),
                    ";".join(
                        f"{srv.service_id}|{srv.name}|{srv.category.value}|{srv.unit_price}|"
                        f"{srv.quantity}|{srv.applied_date.isoformat() if srv.applied_date else ''}"
                        for srv in booking.services
                    ),
                ]
            )

    def _write_services_catalog(self, workbook: Workbook, system: HotelSystem) -> None:
        """Persist the reusable service catalogue."""
        sheet = workbook.create_sheet("services_catalog")
        sheet.append(["service_id", "name", "category", "unit_price", "quantity"])
        for service in system.services.values():
            sheet.append(
                [
                    service.service_id,
                    service.name,
                    service.category.value,
                    service.unit_price,
                    service.quantity,
                ]
            )

    def _read_rooms(self, workbook: Workbook, system: HotelSystem) -> None:
        """Rebuild room objects from the rooms worksheet."""
        if "rooms" not in workbook.sheetnames:
            return
        sheet = workbook["rooms"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            room = Room(
                room_id=str(row[0]),
                room_number=str(row[1]),
                room_type=RoomType(str(row[2])),
                floor=int(row[3]),
                capacity=int(row[4]),
                base_nightly_rate=float(row[5]),
                current_status=RoomStatus(str(row[6])),
                seasonal_multiplier=float(row[7]),
                amenities=[item for item in str(row[8]).split(",") if item],
            )
            system.add_room(room)

    def _read_users(self, workbook: Workbook, system: HotelSystem) -> None:
        """Rebuild users as Guest, Staff, or SuperAdmin based on stored role."""
        if "users" not in workbook.sheetnames:
            return
        sheet = workbook["users"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            role = UserRole(str(row[5]))
            access_level_raw = row[8]
            access_level = int(access_level_raw) if access_level_raw not in (None, "") else 1
            user = build_user_from_role(
                role=role,
                user_id=str(row[0]),
                full_name=str(row[1]),
                email=str(row[2]),
                password_hash=str(row[3]),
                phone=str(row[4]),
                access_level=access_level,
            )
            user.created_at = datetime.fromisoformat(str(row[6]))
            user.is_active = bool(int(row[7]))
            loyalty_points_raw = row[9]
            if isinstance(user, Guest) and loyalty_points_raw not in (None, ""):
                user.loyalty_points = int(loyalty_points_raw)
            if isinstance(user, Staff):
                user.shift = ShiftType.MORNING
            system.add_user(user)

    def _read_bookings(self, workbook: Workbook, system: HotelSystem) -> None:
        """Rebuild bookings and relink each booking ID to the matching guest."""
        if "bookings" not in workbook.sheetnames:
            return
        sheet = workbook["bookings"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            booking = Booking(
                booking_id=str(row[0]),
                guest_id=str(row[1]),
                room_id=str(row[2]),
                check_in_date=date.fromisoformat(str(row[3])),
                check_out_date=date.fromisoformat(str(row[4])),
                status=BookingStatus(str(row[5])),
                total_charges=float(row[6]),
                deposit_paid=float(row[7]),
                notes=str(row[8]) if row[8] is not None else "",
                created_by=str(row[9]) if row[9] is not None else "",
                modified_at=datetime.fromisoformat(str(row[10])),
                services=self._parse_services(str(row[11]) if row[11] is not None else ""),
            )
            system.bookings[booking.booking_id] = booking

            guest = system.users.get(booking.guest_id)
            if isinstance(guest, Guest) and booking.booking_id not in guest.bookings:
                guest.bookings.append(booking.booking_id)

    def _read_services_catalog(self, workbook: Workbook, system: HotelSystem) -> None:
        """Rebuild the service catalogue from Excel."""
        if "services_catalog" not in workbook.sheetnames:
            return
        sheet = workbook["services_catalog"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            service = Service(
                service_id=str(row[0]),
                name=str(row[1]),
                category=ServiceCategory(str(row[2])),
                unit_price=float(row[3]),
                quantity=int(row[4]),
            )
            system.add_service_catalog_item(service)

    def _parse_services(self, raw: str) -> list[Service]:
        """Convert serialized booking service strings back into Service objects."""
        if not raw:
            return []
        services = []
        for token in raw.split(";"):
            parts = token.split("|")
            if len(parts) < 6:
                continue
            applied_date = date.fromisoformat(parts[5]) if parts[5] else None
            services.append(
                Service(
                    service_id=parts[0],
                    name=parts[1],
                    category=ServiceCategory(parts[2]),
                    unit_price=float(parts[3]),
                    quantity=int(parts[4]),
                    applied_date=applied_date,
                )
            )
        return services
